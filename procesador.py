import os
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.drawing.image import Image

# 1. Configuración de carpetas de salida
os.makedirs('output', exist_ok=True)
excel_path = 'output/Reporte_Monitoreo_Ambiental.xlsx'
chart_path = 'output/tendencia_parametros.png'

# 2. Límites Normativos de Referencia (ECA Perú - Cat 3: Agua para Riego y Animales)
# pH: 6.5 - 8.5 | Cadmio (Cd): <= 0.01 mg/L | Plomo (Pb): <= 0.05 mg/L
LIMITES = {
    'pH_min': 6.5,
    'pH_max': 8.5,
    'Cadmio_mgL': 0.01,
    'Plomo_mgL': 0.05
}

# ==============================================================================
# 3. Generación de Datos Reales Simulados (Monitoreo mensual en 3 estaciones)
# ==============================================================================
np.random.seed(42)
fechas = pd.date_range(start='2025-01-01', periods=6, freq='ME').strftime('%Y-%m').tolist() * 3
estaciones = ['Estación E-01 (Aguas Arriba)', 'Estación E-02 (Punto de Vertimiento)', 'Estación E-03 (Aguas Abajo)'] * 6
estaciones.sort()

# Generamos los valores numéricos crudos flotantes originalmente
cadmio_valores = np.random.uniform(0.002, 0.015, 18).round(4)

df = pd.DataFrame({
    'Fecha': fechas,
    'Estación': estaciones,
    'pH': np.random.uniform(6.0, 9.0, 18).round(2),
    'Plomo_mgL': np.random.uniform(0.01, 0.07, 18).round(4)
})

# Creamos la columna procesada analíticamente primero (como números reales para cálculos)
df['Cadmio_Procesado'] = np.where(cadmio_valores < 0.004, 0.001, cadmio_valores)

# Creamos la columna de reporte visual convirtiendo los valores menores a 0.004 en el string "<LOD"
df['Cadmio_mgL'] = np.where(cadmio_valores < 0.004, "<LOD", cadmio_valores.astype(str))


# 4. Generación de Gráfico de Tendencias con Matplotlib
plt.style.use('ggplot')
fig, ax = plt.subplots(1, 2, figsize=(12, 5))

# Gráfico de tendencias de Plomo por Estación
for estacion in df['Estación'].unique():
    data_estacion = df[df['Estación'] == estacion]
    ax[0].plot(data_estacion['Fecha'], data_estacion['Plomo_mgL'], marker='o', label=estacion)
ax[0].axhline(y=LIMITES['Plomo_mgL'], color='red', linestyle='--', label='Límite ECA (0.05 mg/L)')
ax[0].set_title('Evolución del Plomo (Pb) por Campaña')
ax[0].set_ylabel('Concentración (mg/L)')
ax[0].tick_params(axis='x', rotation=45)
ax[0].legend(prop={'size': 8})

# Gráfico de cajas para el pH
df.boxplot(column='pH', by='Estación', ax=ax[1], grid=False)
ax[1].axhline(y=LIMITES['pH_max'], color='red', linestyle='--', label='ECA Máx (8.5)')
ax[1].axhline(y=LIMITES['pH_min'], color='blue', linestyle='--', label='ECA Mín (6.5)')
ax[1].set_title('Distribución de pH por Estación')
ax[1].tick_params(axis='x', rotation=15)

plt.tight_layout()
plt.savefig(chart_path, dpi=100)
plt.close()

# 5. Exportación y Formateo Avanzado en Excel con OpenPyXL
wb = Workbook()
ws = wb.active
ws.title = "Datos de Monitoreo"

# Encabezados
headers = ['Fecha', 'Estación', 'pH', 'Cadmio (mg/L)', 'Plomo (mg/L)']
ws.append(headers)

# Estilos de encabezado
header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.font = header_font
    cell.fill = header_fill

# Estilos para alertas normativas
alerta_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
alerta_font = Font(name='Arial', size=10, color='9C0006', bold=True)

# Escribir filas y validar límites en tiempo real
for index, row in df.iterrows():
    row_data = [row['Fecha'], row['Estación'], row['pH'], row['Cadmio_mgL'], row['Plomo_mgL']]
    ws.append(row_data)
    
    current_row = ws.max_row
    
    # Validar pH
    cell_ph = ws.cell(row=current_row, column=3)
    if cell_ph.value < LIMITES['pH_min'] or cell_ph.value > LIMITES['pH_max']:
        cell_ph.fill = alerta_fill
        cell_ph.font = alerta_font
        
    # Validar Plomo
    cell_pb = ws.cell(row=current_row, column=5)
    if cell_pb.value > LIMITES['Plomo_mgL']:
        cell_pb.fill = alerta_fill
        cell_pb.font = alerta_font

# Autoajustar ancho de columnas
for col in ws.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    col_letter = col[0].column_letter
    ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

# 6. Insertar Gráfico generado en una nueva pestaña de Excel
ws_graficos = wb.create_sheet(title="Reporte Gráfico")
ws_graficos.views.sheetView[0].showGridLines = True
ws_graficos.cell(row=2, column=2, value="Dashboard de Control de Calidad e Impacto Ambiental").font = Font(size=14, bold=True, color='1F4E78')

img = Image(chart_path)
ws_graficos.add_image(img, 'B4')

wb.save(excel_path)
print(f"🎉 ¡Reporte generado con éxito en: {excel_path}!")
