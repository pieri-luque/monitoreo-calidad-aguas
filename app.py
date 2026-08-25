import streamlit as st
import pandas as pd
import numpy as np
import io
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

# Importamos la base de datos reglamentaria
from config_ecas import ECA_AGUA, SINONIMOS_PARAMETROS

st.set_page_config(page_title="Dashboard de Vigilancia Ambiental", page_icon="📊", layout="wide")

# Estilización de fondo oscuro/claro para simular paneles operativos
st.title("📊 Panel Operativo de Calidad Hídrica y Cumplimiento Normativo")
st.markdown("---")

# 1. Configuración y Carga en la barra lateral
st.sidebar.header("🎛️ Panel de Configuración")
archivo_cargado = st.sidebar.file_uploader("Sube la matriz de monitoreo (.xlsx)", type=["xlsx"])

# Selector dinámico de Tipo de Agua / Normativa Regulatoria
tipo_norma = st.sidebar.selectbox(
    "Selecciona la normativa de evaluación (ECA):",
    options=list(ECA_AGUA.keys())
)
limites_actuales = ECA_AGUA[tipo_norma]

if archivo_cargado is not None:
    # Cargar datos crudos
    df_raw = pd.read_excel(archivo_cargado)
    
    # --- MOTOR DE RECONOCIMIENTO Y ESTANDARIZACIÓN DE COLUMNAS ---
    df_procesado = df_raw.copy()
    
    # Renombrar columnas dinámicamente usando el diccionario de sinónimos reglamentarios
    nuevos_nombres = {}
    for col in df_procesado.columns:
        col_clean = str(col).strip().lower().replace("_", "").replace(" ", "")
        if col_clean in SINONIMOS_PARAMETROS:
            nuevos_nombres[col] = SINONIMOS_PARAMETROS[col_clean]
            
    df_procesado.rename(columns=nuevos_nombres, inplace=True)
    st.sidebar.success("✅ Columnas homologadas con éxito a los parámetros reglamentarios.")

    # Limpieza analítica de datos mixtos (<LOD) para cálculos del Dashboard
    for col in df_procesado.columns:
        if col in limites_actuales:
            df_procesado[col] = df_procesado[col].astype(str).str.replace('<LOD', '0.001')
            df_procesado[col] = pd.to_numeric(df_procesado[col], errors='coerce').fillna(0.001)

    # ==============================================================================
    # DISEÑO DEL DASHBOARD (Estilo Visual de Referencia)
    # ==============================================================================
    
    # Fila 1: KPI Cards (Métricas del último muestreo registrado)
    st.markdown(f"### 🌡️ Últimos Valores Reportados vs Límites de la Categoría")
    cols_kpi = st.columns(len(limites_actuales))
    
    for i, (parametro, limites) in enumerate(limites_actuales.items()):
        if parametro in df_procesado.columns:
            ultimo_valor = df_procesado[parametro].iloc[-1]
            
            # Verificar si infringe la norma para alertas visuales
            if "max" in limites and ultimo_valor > limites["max"]:
                estado = "🔴 Excede Límite"
                delta_color = "inverse"
                delta_val = f"Max: {limites['max']}"
            elif "min" in limites and (ultimo_valor < limites["min"] or ultimo_valor > limites["max"]):
                estado = "🔴 Fuera de Rango"
                delta_color = "inverse"
                delta_val = f"Rango: {limites['min']}-{limites['max']}"
            else:
                estado = "🟢 Conforme"
                delta_color = "normal"
                delta_val = "Dentro del límite"
                
            cols_kpi[i].metric(
                label=f"{parametro} ({estado})",
                value=f"{ultimo_valor:.3f}" if parametro != "pH" else f"{ultimo_valor:.2f}",
                delta=delta_val,
                delta_color=delta_color
            )
            
    st.markdown("---")
    
    # Fila 2: Gráficos Operacionales Analíticos Interactivos (Plotly)
    col_izq, col_der = st.columns(2)
    
    with col_izq:
        st.markdown("### 📈 Tendencia Temporal de Contaminantes")
        
        # Creamos la lista de opciones válidas
        opciones_grafico = [p for p in limites_actuales.keys() if p in df_procesado.columns]
        
        if opciones_grafico:
            param_grafico = st.selectbox("Selecciona parámetro para graficar:", opciones_grafico)
            
            # Validamos explícitamente que param_grafico no sea None antes de usarlo
            if param_grafico:
                fig_lineas = px.line(
                    df_procesado, x="Fecha de Muestreo", y=param_grafico, color="Estación de Monitoreo",
                    markers=True, template="plotly_white", title=f"Evolución de {param_grafico} en el tiempo"
                )
                
                # Añadir línea del límite reglamentario dinámicamente de forma segura
                if param_grafico in limites_actuales and "max" in limites_actuales[param_grafico]:
                    fig_lineas.add_hline(y=limites_actuales[param_grafico]["max"], line_dash="dash", line_color="red", annotation_text="Límite Máximo ECA")
                    
                st.plotly_chart(fig_lineas, use_container_width=True)
        else:
            st.warning("⚠️ No se encontraron parámetros homologados en este archivo para la normativa seleccionada.")


    with col_der:
        st.markdown("### 🗺️ Distribución Geográfica / Análisis de Variabilidad")
        
        # Creamos una lista de opciones dinámicas para el Boxplot basado en lo que SÍ se homologó
        opciones_box = [p for p in limites_actuales.keys() if p in df_procesado.columns]
        
        if opciones_box:
            # Si el pH está disponible, lo ponemos como predeterminado; si no, toma el primero de la lista
            index_por_defecto = opciones_box.index("pH") if "pH" in opciones_box else 0
            
            param_box = st.selectbox(
                "Selecciona parámetro para analizar dispersión:", 
                opciones_box, 
                index=index_por_defecto,
                key="sb_boxplot" # Añadimos un key único para que no choque con el otro selectbox
            )
            
            if param_box:
                fig_box = px.box(
                    df_procesado, x="Estación de Monitoreo", y=param_box, color="Estación de Monitoreo",
                    template="plotly_white", title=f"Dispersión de {param_box} por Punto de Muestreo"
                )
                
                # Añadir líneas de límites mínimos y máximos dinámicamente según el parámetro seleccionado
                if param_box in limites_actuales:
                    if "max" in limites_actuales[param_box]:
                        fig_box.add_hline(y=limites_actuales[param_box]["max"], line_dash="dot", line_color="red", annotation_text="Límite Máximo")
                    if "min" in limites_actuales[param_box]:
                        fig_box.add_hline(y=limites_actuales[param_box]["min"], line_dash="dot", line_color="blue", annotation_text="Límite Mínimo")
                        
                st.plotly_chart(fig_box, use_container_width=True)
        else:
            st.info("ℹ️ Esperando homologación de parámetros para generar el diagrama de cajas.")


    # ==============================================================================
    # MOTOR DE EXPORTACIÓN CON FORMATO CONDICIONAL INTEGRADO
    # ==============================================================================
    output_excel = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Cumplimiento"
    
    headers = list(df_raw.columns)
    ws.append(headers)
    
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    alerta_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    alerta_font = Font(name='Arial', size=10, color='9C0006', bold=True)
    
    for col_num, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill

    for index, row in df_procesado.iterrows():
        ws.append(list(df_raw.iloc[index]))
        curr_row = ws.max_row
        
        # Validar cada celda según el mapeo reglamentario dinámico
        for col_idx, col_name in enumerate(df_raw.columns, 1):
            param_oficial = nuevos_nombres.get(col_name, None)
            if param_oficial in limites_actuales:
                val = row[param_oficial]
                lim = limites_actuales[param_oficial]
                
                if "max" in lim and val > lim["max"]:
                    ws.cell(row=curr_row, column=col_idx).fill = alerta_fill
                    ws.cell(row=curr_row, column=col_idx).font = alerta_font
                elif "min" in lim and (val < lim["min"] or val > lim["max"]):
                    ws.cell(row=curr_row, column=col_idx).fill = alerta_fill
                    ws.cell(row=curr_row, column=col_idx).font = alerta_font

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 3, 12)
        
    wb.save(output_excel)
    
    st.markdown("### 📥 Descargas de Auditoría")
    st.download_button(
        label="📥 Descargar Reporte Técnico de Validación Normativa (.xlsx)",
        data=output_excel.getvalue(),
        file_name=f"Reporte_Validado_{tipo_norma.split(':')[0].replace(' ', '_')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
else:
    st.warning("👈 Selecciona el Tipo de Agua en el menú e ingresa la plantilla de datos para desplegar los paneles operativos.")

